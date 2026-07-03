# -*- coding: utf-8 -*-
"""
Generates a deliberately messy customer master data Excel file for the
demo project "Bereinigung von Kundenstammdaten".

Injected chaos (all synthetic, fictional companies):
  - duplicates of the same customer with different spelling
    (Müller GmbH / Mueller GmbH / MÜLLER GMBH / Müler GmbH)
  - typos in e-mail addresses (comma instead of dot, missing @, spaces)
  - inconsistent phone formats (+49..., 0351/..., (0351) ..., digits only)
  - empty required fields
  - leading/trailing/double spaces, ALL-CAPS values

Run:  python generate_input.py
"""

import os
import random

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

random.seed(11)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(BASE_DIR, "input")
os.makedirs(INPUT_DIR, exist_ok=True)

FIRMEN = [
    "Müller GmbH", "Schneider & Sohn KG", "Bäckerei Vogt", "Autohaus Lindner GmbH",
    "Elektro Brenner", "Gartenbau Wiese GmbH", "Metallbau Krause AG", "Friseur Salon Chic",
    "Steuerbüro Albrecht", "Malermeister Kunz", "Spedition Falkner GmbH", "Optik Sonnberg",
    "Tischlerei Holzmann", "Reinigung Perle UG", "Apotheke am Markt", "Foto Studio Licht",
    "Kanzlei Weber & Partner", "Sanitär Quelle GmbH", "Buchhandlung Seitenweise",
    "Dachdecker First KG", "Café Morgenrot", "Werbeagentur Pixelhaus GmbH",
    "Fahrschule Startklar", "Immobilien Schlüssel GmbH", "Catering Gaumenfreude",
    "Physiotherapie Balance", "Druck & Papier Held", "IT-Systeme Nordlicht GmbH",
    "Bestattungen Ruhestein", "Modehaus Eleganza", "Schlosserei Eisenhart",
    "Reisebüro Fernweh GmbH", "Zahnarztpraxis Dr. Sommer", "Getränke Quellfrisch KG",
    "Blumen Rosenthal", "Umzüge Packan GmbH", "Feinkost Delikat", "Kfz-Werkstatt Kolben",
    "Musikschule Tonleiter", "Versicherungsbüro Schutz & Co.",
]

VORNAMEN = ["Hans", "Petra", "Jürgen", "Sabine", "Klaus", "Monika", "Dieter",
            "Renate", "Wolfgang", "Ute", "Bernd", "Ingrid", "Frank", "Heike"]
NACHNAMEN = ["Müller", "Schmidt", "Weber", "Fischer", "Meyer", "Wagner", "Becker",
             "Hoffmann", "Schulz", "Koch", "Richter", "Klein", "Wolf", "Neumann"]
STRASSEN = ["Hauptstraße", "Bahnhofstraße", "Gartenweg", "Lindenallee", "Mühlgasse",
            "Am Anger", "Ringstraße", "Feldweg", "Schulstraße", "Marktplatz"]
ORTE = [("01067", "Dresden"), ("01069", "Dresden"), ("01099", "Dresden"),
        ("04103", "Leipzig"), ("04109", "Leipzig"), ("09111", "Chemnitz"),
        ("02625", "Bautzen"), ("01796", "Pirna"), ("01445", "Radebeul")]
VORWAHLEN = ["351", "341", "371", "3591", "3501"]


def slug(firma):
    s = firma.lower()
    for a, b in [("ä", "ae"), ("ö", "oe"), ("ü", "ue"), ("ß", "ss")]:
        s = s.replace(a, b)
    s = "".join(ch for ch in s if ch.isalnum() or ch == " ")
    return s.split()[0]


def make_kunde(firma):
    person = f"{random.choice(VORNAMEN)} {random.choice(NACHNAMEN)}"
    strasse = f"{random.choice(STRASSEN)} {random.randint(1, 120)}"
    plz, ort = random.choice(ORTE)
    vorwahl = random.choice(VORWAHLEN)
    nummer = random.randint(200000, 999999)
    telefon = f"+49 {vorwahl} {nummer}"
    email = f"info@{slug(firma)}.example"
    return [firma, person, strasse, plz, ort, telefon, email]


def variiere_telefon(tel):
    """same number, different notation"""
    teile = tel.replace("+49 ", "").split(" ")
    vorwahl, nummer = teile[0], teile[1]
    varianten = [
        f"0{vorwahl}/{nummer}",
        f"(0{vorwahl}) {nummer}",
        f"0{vorwahl} - {nummer}",
        f"0049{vorwahl}{nummer}",
        f"0{vorwahl}{nummer}",
    ]
    return random.choice(varianten)


def kaputte_email(email):
    art = random.randint(0, 3)
    if art == 0:
        return email.replace(".example", ",example")      # Komma statt Punkt
    if art == 1:
        return email.replace("@", "")                     # @ fehlt
    if art == 2:
        return email.replace("@", " @ ")                  # Leerzeichen
    return email.upper()                                  # Grossschreibung


def umlaut_variante(firma):
    for a, b in [("ü", "ue"), ("ö", "oe"), ("ä", "ae"), ("ß", "ss")]:
        if a in firma:
            return firma.replace(a, b)
    return firma


def tippfehler(firma):
    """drop one letter in the first word"""
    wort = firma.split()[0]
    if len(wort) > 4:
        pos = random.randint(2, len(wort) - 2)
        return firma.replace(wort, wort[:pos] + wort[pos + 1:], 1)
    return firma


# ------------------------------------------------------------ base records
basis = [make_kunde(f) for f in FIRMEN]

zeilen = []
for kunde in basis:
    zeilen.append(list(kunde))

# ------------------------------------------------------------ duplicates
# pick 6 customers and add 1-2 variants of each
dupl_quellen = random.sample(range(len(basis)), 6)
for idx in dupl_quellen:
    orig = basis[idx]
    for _ in range(random.randint(1, 2)):
        d = list(orig)
        art = random.randint(0, 3)
        if art == 0:
            d[0] = umlaut_variante(d[0])            # Mueller statt Müller
        elif art == 1:
            d[0] = d[0].upper()                     # MÜLLER GMBH
        elif art == 2:
            d[0] = tippfehler(d[0])                 # Müler GmbH
        else:
            d[0] = d[0].lower()                     # müller gmbh
        # duplicates often carry diverging secondary data
        if random.random() < 0.5:
            d[5] = variiere_telefon(d[5])
        if random.random() < 0.4:
            d[6] = ""                               # fehlende E-Mail im Duplikat
        zeilen.append(d)

# ------------------------------------------------------------ other chaos
random.shuffle(zeilen)

for i, z in enumerate(zeilen):
    r = random.random()
    if r < 0.25:
        z[5] = variiere_telefon(z[5]) if z[5] else z[5]     # Telefonformat
    if 0.25 <= r < 0.40:
        z[6] = kaputte_email(z[6]) if z[6] else z[6]        # kaputte E-Mail
    if 0.40 <= r < 0.50:
        z[random.choice([1, 5, 6])] = ""                    # Pflichtfeld leer
    if 0.50 <= r < 0.60:
        z[0] = "  " + z[0].replace(" ", "  ", 1) + " "      # Leerzeichen-Chaos
    if 0.60 <= r < 0.68:
        z[4] = z[4].upper()                                 # ORT in Grossbuchstaben
    if 0.68 <= r < 0.74:
        z[3] = ""                                           # PLZ fehlt

# ------------------------------------------------------------ write Excel
wb = Workbook()
ws = wb.active
ws.title = "Kunden"

headers = ["Firma", "Ansprechpartner", "Straße", "PLZ", "Ort", "Telefon", "E-Mail"]
fill = PatternFill("solid", fgColor="1F4E79")
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = fill

for r, z in enumerate(zeilen, 2):
    for col, wert in enumerate(z, 1):
        ws.cell(row=r, column=col, value=wert)

for col, breite in zip("ABCDEFG", [32, 20, 20, 8, 14, 22, 34]):
    ws.column_dimensions[col].width = breite

pfad = os.path.join(INPUT_DIR, "Kundenliste_roh.xlsx")
wb.save(pfad)
print(f"OK: {pfad} ({len(zeilen)} Zeilen, davon {len(zeilen) - len(basis)} Duplikate)")
